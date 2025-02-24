from selenium import webdriver
from selenium.webdriver.safari.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pytesseract
from PIL import Image
import time
import pandas as pd
import platform
import urllib3
import requests
import pytesseract
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import numpy as np

# 禁用SSL警告
urllib3.disable_warnings()

# 设置Safari WebDriver
driver = webdriver.Safari()
driver.maximize_window()  # 添加这行来最大化窗口
driver.get("https://gdgpo.czt.gd.gov.cn/cms-gd/site/guangdong/cggg/index.html")

# 等待页面完全加载
time.sleep(2)  # 给予足够的加载时间

# 等待页面加载
wait = WebDriverWait(driver, 2)

# 处理验证码
# 添加区域选择函数
def select_area(driver, wait, area_code, area_name):
    try:
        # 点击地区选择框
        area_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.layui-treeSelect .layui-select-title")))
        area_input.click()
        time.sleep(2)
        
        # 选择指定地区
        area_option = wait.until(EC.element_to_be_clickable((By.XPATH, f"//li[contains(@code, '{area_code}')]//span[contains(@class, 'node_name')]")))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", area_option)
        time.sleep(1)
        
        driver.execute_script("arguments[0].click();", area_option)
        time.sleep(2)
        
        print(f"已选择{area_name}")
        return area_name
    except Exception as e:
        print(f"选择{area_name}失败:", e)
        return "未知地区"

# 优化验证码处理函数
def handle_captcha(driver, wait):
    def process_image(img):
        img_array = np.array(img)
        red_channel = img_array[:, :, 0]
        green_channel = img_array[:, :, 1]
        blue_channel = img_array[:, :, 2]
        red_mask = (red_channel > 150) & (green_channel < 100) & (blue_channel < 100)
        binary_image = np.zeros((img_array.shape[0], img_array.shape[1]), dtype=np.uint8)
        binary_image[~red_mask] = 255
        return Image.fromarray(binary_image)

    def verify_code(code):
        if not code or len(code) != 4 or not code.isdigit():
            return False
        return True

    max_retries = 4
    for retry in range(max_retries):
        try:
            # 获取和处理验证码图片
            img_element = wait.until(EC.presence_of_element_located((By.ID, "code_img")))
            driver.execute_script("arguments[0].scrollIntoView();", img_element)
            time.sleep(2)
            
            screenshot = Image.open(BytesIO(driver.get_screenshot_as_png()))
            location = img_element.location
            size = img_element.size
            captcha_image = screenshot.crop((
                location['x'], location['y'],
                location['x'] + size['width'],
                location['y'] + size['height']
            ))
            
            processed_image = process_image(captcha_image)
            code = pytesseract.image_to_string(
                processed_image,
                config='--psm 7 -c tessedit_char_whitelist=0123456789'
            ).strip()
            
            if not verify_code(code):
                print(f"验证码识别结果异常: {code}，重试...")
                continue
            
            print(f"识别到验证码: {code}")
            
            # 输入验证码并提交
            code_input = wait.until(EC.presence_of_element_located((By.ID, "verifycode")))
            code_input.clear()
            code_input.send_keys(code)
            
            search_button = wait.until(EC.element_to_be_clickable((By.ID, "Inquire")))
            driver.execute_script("arguments[0].click();", search_button)
            time.sleep(3)
            
            # 验证结果
            list_container = driver.find_element(By.CLASS_NAME, "procurementAnnouncementShowList")
            if list_container.find_elements(By.TAG_NAME, "li"):
                print("验证码正确，列表已加载")
                return True
            
        except Exception as e:
            print(f"验证码处理失败: {e}")
        
        time.sleep(2)
    return False

# 数据获取优化
def fetch_page_data(driver, wait, page):
    try:
        if page > 1:
            pagination = wait.until(EC.presence_of_element_located((By.ID, "pagination")))
            next_page = wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//div[@id='pagination']//li[text()='{page}']")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_page)
            driver.execute_script("arguments[0].click();", next_page)
            time.sleep(3)
        
        list_container = wait.until(EC.presence_of_element_located((By.ID, "procurementAnnouncementShowList")))
        items = list_container.find_elements(By.TAG_NAME, "li")
        
        page_data = []
        for item in items:
            try:
                link = item.find_element(By.TAG_NAME, "a")
                spans = link.find_elements(By.TAG_NAME, "span")
                page_data.append({
                    "标题": spans[0].text,
                    "地区": spans[1].text,
                    "发布日期": spans[2].text,
                    "网址": link.get_attribute("href")
                })
            except Exception as e:
                print(f"提取列表项信息失败: {e}")
        
        return page_data
    except Exception as e:
        print(f"获取第{page}页数据失败: {e}")
        return []

# 在主代码中使用这些函数
# 主流程
try:
    # 选择地区
    area_i = select_area(driver, wait, '440601', '佛山市')  # 或其他地区代码和名称 广州440101 广东省440001
    
    # 选择采购公告
    notice_type = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@codes='00102']")))#00102中标（成交）结果公告 00101采购公告
    notice_type.click()
    time.sleep(1)

    # 选择服务类别
    service_type = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@codes='3']")))
    service_type.click()
    time.sleep(1)

    # 清空并设置年份输入框
    year_input = wait.until(EC.presence_of_element_located((By.ID, "stateDate")))
    year_input.clear()
    #year_input.send_keys("2024")
    time.sleep(4)

    # 设置时间范围
    time_range = wait.until(EC.presence_of_element_located((By.ID, "uploadTime")))
    driver.execute_script("arguments[0].scrollIntoView();", time_range)
    time.sleep(2)
    

    
    # 设置详细时间范围
    driver.execute_script(
        "arguments[0].value = '2022-01-01 00:00:00 至 2022-12-31 00:00:00'",
        time_range
    )
    driver.execute_script("arguments[0].dispatchEvent(new Event('change'))", time_range)
    time.sleep(1)

    # 验证码处理
    if not handle_captcha(driver, wait):
        print("验证码处理失败,重试...")
        driver.refresh()
        time.sleep(2)
        if not handle_captcha(driver, wait):
            print("验证码处理失败次数过多,退出")
            exit(1)

    # 获取数据
    all_data = []
    for page in range(1, 260):
        page_data = fetch_page_data(driver, wait, page)
        if not page_data:
            break
        all_data.extend(page_data)
        print(f"第 {page} 页数据获取完成")

    # 保存数据到Excel
    from datetime import datetime
    today_date = datetime.today().strftime('%Y-%m-%d_%H%M%S')
    df = pd.DataFrame(all_data)
    df.index = range(1, len(df) + 1)
    df.index.name = "序号"
    excel_filename = f"中标公告列表_{area_i}_{today_date}.xlsx"
    df.to_excel(excel_filename, index=True)
    print(f"\n已保存全部 {len(all_data)} 条数据到 {excel_filename}")

    # 调用格式化函数
    try:
        # 在主流程之前添加 format_excel 函数定义
        def format_excel(filename):
            wb = load_workbook(filename)
            ws = wb.active
            
            # 设置单元格边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置所有单元格的格式
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(name='宋体', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
            
            # 设置表头格式（淡蓝色填充）
            header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # 淡蓝色
            header_font = Font(name='宋体', size=10, bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # 设置列宽
            column_widths = {
                1: 10,    # 序号
                2: 20,    # 标题
                3: 10,    # 地区
                4: 10,    # 发布日期
                5: 30,    # 网址
            }
            
            for col_num, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col_num)].width = width
            
            # 设置行高
            for row in ws.rows:
                ws.row_dimensions[row[0].row].height = 30
            
            wb.save(filename)
        print("Excel文件格式化完成")
    except Exception as e:
        print(f"Excel格式化失败: {e}")

except Exception as e:
    print(f"程序执行失败: {str(e)}")

finally:
    driver.quit()
