import re

import requests
from bs4 import BeautifulSoup
import pandas as pd
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry

# 定义要抓取的URL列表
urls = [
    "https://gdgpo.czt.gd.gov.cn/freecms/site/gd/ggxx/info/2025/8a7e7b0294cce056019527a5af9559b5.html?noticeType=001025",
    
    "https://gdgpo.czt.gd.gov.cn/freecms/site/gd/ggxx/info/2025/8a7eb0b494ccd3870195223decdd77bc.html?noticeType=001021"
]

# 创建一个会话
session = requests.Session()

# 设置重试策略
retries = Retry(total=5, backoff_factor=1, status_forcelist=[502, 503, 504])
session.mount('http://', HTTPAdapter(max_retries=retries))
session.mount('https://', HTTPAdapter(max_retries=retries))

# 设置请求头
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}

# 创建一个列表来存储所有数据
all_data = []

# 目标网页URL
url = "https://gdgpo.czt.gd.gov.cn/freecms/site/gd/ggxx/info/2025/8a7e970894ccd1b001951d44cca700ca.html?noticeType=001011"

# 发送HTTP请求获取网页内容，设置超时
response = session.get(url, headers=headers, timeout=10)
response.encoding = 'utf-8'
html_content = response.text

# 使用BeautifulSoup解析网页内容
soup = BeautifulSoup(html_content, 'html.parser')

def convert_to_wan(money_str):
    if not money_str:
        return None
    # 移除"元"和逗号，转换为浮点数
    clean_number = money_str.replace('元', '').replace(',', '').strip()
    return "{:.2f}万元".format(float(clean_number)/10000)

def convert_to_wan2(money_str):
    if not money_str:
        return None
    # 移除"预算金额："，转换为浮点数
    clean_number = money_str.replace('预算金额：', '').strip()
    return "{:.2f}万元".format(float(clean_number)/10000)

# 遍历每个URL并抓取内容
for url in urls:
    response = session.get(url, headers=headers, timeout=10)
    response.encoding = 'utf-8'
    html_content = response.text

    # 使用BeautifulSoup解析网页内容
    soup = BeautifulSoup(html_content, 'html.parser')

    # 抓取所需信息
    data = {
        "项目名称": soup.find('h4', string=re.compile('项目名称')).text.split('：')[1].strip() ,
        "中标公示日期": soup.find('span', id='f_noticeTime').text.replace('发布时间：', '').strip().split()[0] ,
       
        "预计到期时间":  soup.find('p', string=re.compile('人数')).find('span', class_='u-content').text.strip() if soup.find('p', string=re.compile('人数')) else None,
        #"项目编号": soup.find('p', string=re.compile('项目编号')).text.split('：')[1].strip() ,
        "采购单位": soup.find('span', id='_notice_content_noticePurchase-purchaserOrgName').text,
        "招标代理单位": soup.find('span', id='_notice_content_noticeAgency-agencyName').text ,
       # "招标方式": soup.find('p', string=re.compile('采购方式')).text.split('：')[1].strip() ,
       # "服务类型": soup.find('span', id='f_catalogueNameList').text.replace('采购品目：', '').strip() ,
       # "投标形式": '电子标 ' + soup.find('span', id='_notice_content_noticeBidTime-bidFileSubmitAddress').text.replace('https://gdgpo.czt.gd.gov.cn/', '').strip(),
       # "招标预算金额": "{:.2f}万元".format(float(soup.find('span', id="f_budget").text.replace('预算金额：', '').strip())/10000) if soup.find('span', id="f_budget") and soup.find('span', id="f_budget").text else None,
       # 用另外一种方式： 
        "招标预算金额": (
            convert_to_wan2(soup.find('span', id="f_budget").text)
            if soup.find('span', id="f_budget") and soup.find('span', id="f_budget").text
            else None
        ),

        "中标金额": (
            convert_to_wan(soup.find('td', class_='alignright').find('span').text)
            if soup.find('td', class_='alignright') and soup.find('td', class_='alignright').find('span')
            else None
        ),
        #"合同履行期限": soup.find('p', string=re.compile('合同履行期限')).text.split('：')[1].strip() ,
        # 查找包含"合同履行期限"文本的<p>标签,如果找到则继续查找其中的class='u-content'的<span>标签并获取文本内容,
        # 如果未找到包含"合同履行期限"的<p>标签则返回None
        "服务时间": (
            soup.find('div', class_='noticeBidResult-noticeBidResult _notice_content_noticeBidResult-noticeBidResult supplierDetail dynamic-form-editor').find_all('tr')[1].find_all('td')[5].text.strip() if soup.find('div', class_='noticeBidResult-noticeBidResult _notice_content_noticeBidResult-noticeBidResult supplierDetail dynamic-form-editor') else None
        ),
        "中标单位": soup.find('td', style="width:400px;word-break:break-all;").text.strip(),
        #"人数": soup.find('p', string=re.compile('人数')).find('span', class_='u-content').text.strip() if soup.find('p', string=re.compile('人数')) else None,
        #"项目隐形成本": soup.find('p', string=re.compile('资格条件')).find('span', class_='u-content').text.strip() if soup.find('p', string=re.compile('资格条件')) else None,
        "备注": url
        
    }
    
    # 将数据添加到列表中
    all_data.append(data)

# 将所有数据转换为DataFrame
df = pd.DataFrame(all_data)

# 生成Excel表格
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建Excel文件
today_date = datetime.today().strftime('%Y-%m-%d_%H%M%S')
file_name = r'招标信息_{}.xlsx'.format(today_date)
df.index = range(1, len(df) + 1)  # 添加从1开始的序号
df.index.name = '序号'  # 设置序号列的标题
df.to_excel(file_name, index=True)  # 将index设为True以显示序号列

def format_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    
    # 设置单元格边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置所有单元格的格式
    for row in ws.iter_rows():
        for cell in row:
            # 设置字体
            cell.font = Font(name='宋体', size=10)
            # 设置对齐方式
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 设置边框
            cell.border = thin_border
    
    # 设置表头格式
    header_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    header_font = Font(name='宋体', size=10, bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    # 设置列宽
    column_widths = {
        1: 5.2,   # 序号
        2: 41,    # 项目名称
        3: 11.5,  # 公示时间
        4: 12,  # 预计到期日期
        5: 20,    # 采购单位
        6: 25,    # 招标代理单位
        7: 12.5,    # 招标预算金额
        8: 12,    # 中标金额
        9: 25,    # 服务时间
        10: 35,   # 中标单位
        11: 20,   # 备注
        12: 15,   # 服务期
        13: 40,   # 岗位设置/工作内容
        14: 8,    # 人数
        15: 40,   # 项目隐形成本
        16: 50,   # 备注
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # 设置行高
    for row in ws.rows:
        ws.row_dimensions[row[0].row].height = 30
        
    wb.save(filename)

# 调用格式化函数
format_excel(file_name)
